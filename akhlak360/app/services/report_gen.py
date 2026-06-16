"""
app/services/report_gen.py — Generate Laporan PDF dan Excel
Sistem Penilaian 360° Core Values AKHLAK
"""

import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))))
from database import query_db, execute_db
from datetime import date, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def get_report_data(id_period: int) -> dict:
    """Ambil semua data yang dibutuhkan untuk laporan."""
    periode = query_db("SELECT * FROM assessment_periods WHERE id_period=?", (id_period,), one=True)
    if not periode:
        return None

    hasil_list = query_db(
        """SELECT h.*, e.nama, e.nip, e.division, e.position
           FROM hasil_akhirs h
           JOIN employees e ON h.id_employee = e.user_id
           WHERE h.id_period = ?
           ORDER BY e.division, h.total_score DESC""",
        (id_period,)
    )

    # Statistik per divisi
    division_stats = query_db(
        """SELECT e.division,
                  COUNT(h.id_result) as jumlah,
                  AVG(h.total_score) as avg_score,
                  MAX(h.total_score) as max_score,
                  MIN(h.total_score) as min_score
           FROM hasil_akhirs h
           JOIN employees e ON h.id_employee = e.user_id
           WHERE h.id_period = ?
           GROUP BY e.division""",
        (id_period,)
    )

    return {
        'periode': periode,
        'hasil_list': hasil_list,
        'division_stats': division_stats,
        'generated_at': datetime.now().strftime('%d %B %Y %H:%M'),
    }


def generate_excel(id_period: int, generated_by: int = None) -> str:
    """
    Generate laporan Excel untuk satu periode.

    Returns:
        str: path file Excel yang dihasilkan
    """
    data = get_report_data(id_period)
    if not data:
        raise ValueError(f"Periode {id_period} tidak ditemukan.")

    # Setup workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan AKHLAK 360°"

    # ── Styling helpers ──
    hijau_tua  = PatternFill("solid", fgColor="006B3C")
    hijau_muda = PatternFill("solid", fgColor="00A86B")
    kuning     = PatternFill("solid", fgColor="F5A623")
    abu        = PatternFill("solid", fgColor="F7F9FC")
    putih_font = Font(name='Calibri', color='FFFFFF', bold=True, size=12)
    header_font = Font(name='Calibri', bold=True, size=11)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    row = 1

    # ── Header Laporan ──
    ws.merge_cells(f'A{row}:J{row}')
    ws[f'A{row}'] = 'PT. ENERGI NUSANTARA'
    ws[f'A{row}'].font = Font(name='Calibri', bold=True, size=16, color='FFFFFF')
    ws[f'A{row}'].fill = hijau_tua
    ws[f'A{row}'].alignment = center
    row += 1

    ws.merge_cells(f'A{row}:J{row}')
    ws[f'A{row}'] = 'LAPORAN PENILAIAN 360° CORE VALUES AKHLAK'
    ws[f'A{row}'].font = putih_font
    ws[f'A{row}'].fill = hijau_tua
    ws[f'A{row}'].alignment = center
    row += 1

    ws.merge_cells(f'A{row}:J{row}')
    ws[f'A{row}'] = f'Periode: {data["periode"]["period_name"]} | Dihasilkan: {data["generated_at"]}'
    ws[f'A{row}'].font = Font(name='Calibri', size=10, color='FFFFFF')
    ws[f'A{row}'].fill = hijau_muda
    ws[f'A{row}'].alignment = center
    row += 2

    # ── Header Tabel ──
    headers = ['No', 'Nama Karyawan', 'NIP', 'Divisi', 'Jabatan',
               'Skor Atasan', 'Skor Bawahan', 'Skor Rekan', 'Skor Self', 'Total Skor', 'Kategori']
    # Tambahkan kolom 11
    ws.merge_cells(f'A{row}:A{row}')  # reset merge jika ada

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = putih_font
        cell.fill = hijau_tua
        cell.alignment = center
        cell.border = thin_border
    row += 1

    # ── Data Karyawan ──
    for i, h in enumerate(data['hasil_list'], start=1):
        values = [
            i,
            h['nama'],
            h['nip'],
            h['division'],
            h['position'],
            h['score_atasan'] or '-',
            h['score_bawahan'] or '-',
            h['score_rekan'] or '-',
            h['score_self'] or '-',
            h['total_score'] or '-',
            h['category'] or '-',
        ]
        fill = abu if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1

    row += 1

    # ── Summary per Divisi ──
    ws.merge_cells(f'A{row}:K{row}')
    ws[f'A{row}'] = 'RINGKASAN PER DIVISI'
    ws[f'A{row}'].font = putih_font
    ws[f'A{row}'].fill = kuning
    ws[f'A{row}'].alignment = center
    row += 1

    div_headers = ['Divisi', 'Jumlah Karyawan', 'Rata-rata Skor', 'Skor Tertinggi', 'Skor Terendah']
    for col_idx, header in enumerate(div_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = hijau_muda
        cell.alignment = center
        cell.border = thin_border
    row += 1

    for ds in data['division_stats']:
        values = [
            ds['division'],
            ds['jumlah'],
            round(ds['avg_score'], 2) if ds['avg_score'] else '-',
            round(ds['max_score'], 2) if ds['max_score'] else '-',
            round(ds['min_score'], 2) if ds['min_score'] else '-',
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        row += 1

    # ── Atur lebar kolom ──
    col_widths = [5, 25, 15, 18, 20, 13, 13, 13, 12, 12, 15]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Atur tinggi baris header
    ws.row_dimensions[4].height = 25

    # ── Simpan file ──
    from config import config
    os.makedirs(config.REPORTS_FOLDER, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    period_name_clean = data['periode']['period_name'].replace(' ', '_')
    filename = f'Laporan_AKHLAK_{period_name_clean}_{timestamp}.xlsx'
    file_path = os.path.join(config.REPORTS_FOLDER, filename)
    wb.save(file_path)

    # Log ke database
    execute_db(
        "INSERT INTO hasil_reports (id_period, report_type, generate_date, generated_by, file_path) VALUES (?,?,?,?,?)",
        (id_period, 'excel', date.today().isoformat(), generated_by, file_path)
    )

    return file_path


def generate_pdf(id_period: int, generated_by: int = None) -> str:
    """
    Generate laporan PDF menggunakan WeasyPrint.

    Returns:
        str: path file PDF yang dihasilkan
    """
    data = get_report_data(id_period)
    if not data:
        raise ValueError(f"Periode {id_period} tidak ditemukan.")

    # Buat HTML laporan
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <style>
            body {{ font-family: Arial, sans-serif; font-size: 10pt; color: #1A2332; margin: 2cm; }}
            h1 {{ color: #006B3C; text-align: center; font-size: 16pt; }}
            h2 {{ color: #00A86B; font-size: 12pt; }}
            .header {{ background: #006B3C; color: white; text-align: center; padding: 15px; border-radius: 8px; }}
            .header h1 {{ color: white; margin: 0; }}
            .header p  {{ margin: 5px 0 0; font-size: 10pt; }}
            table {{ width: 100%; border-collapse: collapse; margin: 15px 0; }}
            th {{ background: #006B3C; color: white; padding: 8px; text-align: center; font-size: 9pt; }}
            td {{ border: 1px solid #ddd; padding: 6px 8px; font-size: 9pt; }}
            tr:nth-child(even) {{ background: #F7F9FC; }}
            .badge-sangat-baik {{ background: #006B3C; color: white; padding: 2px 8px; border-radius: 12px; }}
            .badge-baik {{ background: #00A86B; color: white; padding: 2px 8px; border-radius: 12px; }}
            .badge-cukup {{ background: #F5A623; color: white; padding: 2px 8px; border-radius: 12px; }}
            .badge-kurang {{ background: #E63946; color: white; padding: 2px 8px; border-radius: 12px; }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1>PT. ENERGI NUSANTARA</h1>
            <p>Laporan Penilaian 360° Core Values AKHLAK</p>
            <p>Periode: {data['periode']['period_name']} | Dihasilkan: {data['generated_at']}</p>
        </div>

        <h2>Hasil Penilaian Karyawan</h2>
        <table>
            <tr>
                <th>No</th><th>Nama</th><th>NIP</th><th>Divisi</th>
                <th>Atasan</th><th>Bawahan</th><th>Rekan</th><th>Self</th>
                <th>Total</th><th>Kategori</th>
            </tr>
    """

    for i, h in enumerate(data['hasil_list'], 1):
        cat = h['category'] or '-'
        badge_class = 'badge-' + cat.lower().replace(' ', '-')
        html_content += f"""
            <tr>
                <td style="text-align:center">{i}</td>
                <td>{h['nama']}</td>
                <td style="text-align:center">{h['nip']}</td>
                <td>{h['division']}</td>
                <td style="text-align:center">{h['score_atasan'] or '-'}</td>
                <td style="text-align:center">{h['score_bawahan'] or '-'}</td>
                <td style="text-align:center">{h['score_rekan'] or '-'}</td>
                <td style="text-align:center">{h['score_self'] or '-'}</td>
                <td style="text-align:center"><strong>{h['total_score'] or '-'}</strong></td>
                <td style="text-align:center"><span class="{badge_class}">{cat}</span></td>
            </tr>
        """

    html_content += """
        </table>

        <h2>Ringkasan per Divisi</h2>
        <table>
            <tr>
                <th>Divisi</th><th>Jumlah Karyawan</th>
                <th>Rata-rata Skor</th><th>Tertinggi</th><th>Terendah</th>
            </tr>
    """

    for ds in data['division_stats']:
        html_content += f"""
            <tr>
                <td>{ds['division']}</td>
                <td style="text-align:center">{ds['jumlah']}</td>
                <td style="text-align:center">{round(ds['avg_score'], 2) if ds['avg_score'] else '-'}</td>
                <td style="text-align:center">{round(ds['max_score'], 2) if ds['max_score'] else '-'}</td>
                <td style="text-align:center">{round(ds['min_score'], 2) if ds['min_score'] else '-'}</td>
            </tr>
        """

    html_content += """
        </table>
    </body>
    </html>
    """

    # Simpan file PDF menggunakan WeasyPrint
    from config import config
    os.makedirs(config.REPORTS_FOLDER, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    period_name_clean = data['periode']['period_name'].replace(' ', '_')
    filename = f'Laporan_AKHLAK_{period_name_clean}_{timestamp}.pdf'
    file_path = os.path.join(config.REPORTS_FOLDER, filename)

    try:
        from weasyprint import HTML
        HTML(string=html_content).write_pdf(file_path)
    except ImportError:
        # Fallback: simpan sebagai HTML jika WeasyPrint tidak tersedia
        file_path = file_path.replace('.pdf', '.html')
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(html_content)

    # Log ke database
    execute_db(
        "INSERT INTO hasil_reports (id_period, report_type, generate_date, generated_by, file_path) VALUES (?,?,?,?,?)",
        (id_period, 'pdf', date.today().isoformat(), generated_by, file_path)
    )

    return file_path
